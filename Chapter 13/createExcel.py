# import openpyxl
# wb = openpyxl.Workbook() #Create a blank workbook 
# wb.sheetnames
# sheet = wb.active
# sheet.title

# sheet.title =  'Spam Bacon Eggs Sheet' #change title
# print(wb.sheetnames)



import openpyxl
wb = openpyxl.Workbook()

sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['C5'] = 'Two merged cells'

wb.save('merged.xlsx')



